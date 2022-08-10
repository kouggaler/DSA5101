def getlines(path):
    file = open(path, 'r')

    Lines = file.readlines() 
    Lines2 = []
    for line in Lines:
        line = line.strip()
        if line != "":
            Lines2.append(line)
    return Lines2

def getcountdict(Lines):
    words = []
    for line in Lines:
        words.extend(word for word in line.split(' '))

    punc = '''!()-[]{};:'"\,<>./?@#$%^&*_~'''
    numbers = ['0','1','2','3','4','5','6','7','8','9']
    words2 = []
    for word in words:
        for i in word:
            if i in punc:
                word = word.replace(i,"")
            word = word.lower()
        if all(num not in word for num in numbers):
            words2.append(word) 

    set_word = set(words2)
    countdict = {}
    for unique in set_word:
        countdict2 = {unique : words2.count(unique)}
        countdict.update(countdict2)

    return countdict

pathA = 'fileA.txt'
pathB = 'fileB.txt'

LinesA = getlines(pathA)
LinesB = getlines(pathB)

countdictA = getcountdict(LinesA)
countdictB = getcountdict(LinesB)

finalcountdict = {}
for kA,vA in countdictA.items():
    if kA not in list(countdictB.keys()):
        finalcountdict2 = {kA: [vA, 0]}
        finalcountdict.update(finalcountdict2)
    if kA in list(countdictB.keys()):
        if vA > countdictB[kA]:
            finalcountdict2 = {kA: [vA, countdictB[kA]]}
            finalcountdict.update(finalcountdict2)

sortedcounttuple = sorted(finalcountdict.items(),key=lambda item:item[1], reverse = True)

import openpyxl as xl
from openpyxl.chart import BarChart, Reference
from openpyxl import Workbook
wb = Workbook()
ws = wb.active
ws.title = 'Q3'
ws.append(['Word','#(Occurrences) in A','#(Occurrences) in B'])

sheet = wb['Q3']

for i in range (len(sortedcounttuple)):
    cell_word = sheet.cell(i+2,1)
    cell_occA = sheet.cell(i+2,2)
    cell_occB = sheet.cell(i+2,3)
    cell_word.value = sortedcounttuple[i][0]
    cell_occA.value = sortedcounttuple[i][1][0]
    cell_occB.value = sortedcounttuple[i][1][1]

data = Reference(sheet, min_row = 1,
                  max_row = min(6,sheet.max_row),
                  min_col = 2,
                  max_col = 3)
cats = Reference(sheet, min_col = 1, min_row = 2,  max_row = min(6,sheet.max_row))
    
chart = BarChart()
chart.type = "col"
chart.style = 10
chart.title = "Top 5 words in fileA"
chart.y_axis.title = "Occurances"
chart.x_axis.title = "Words"

chart.add_data(data, titles_from_data = True)
chart.set_categories(cats)
chart.shape = 5
sheet.add_chart(chart,'G2')

wb.save('WordOccurence.xlsx')